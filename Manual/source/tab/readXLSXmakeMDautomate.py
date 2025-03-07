from openpyxl import load_workbook
from markdown import markdown
import re
import os

def excel_to_markdown_table(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = True

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = True
            else:
                row_data.append(cell.value)
                if make_bold:
                    row_data[0] = "**" + row_data[0] + "**"
                    make_bold = False  # Reset bold flag after making first cell bold
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            consecutive_blank_rows = 0  # Reset consecutive_blank_rows counter
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table
    

def nonBold(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = False

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = False
            else:
                row_data.append(cell.value)
                if make_bold:
                    row_data[0] = "**" + row_data[0] + "**"
                    make_bold = False  # Reset bold flag after making first cell bold
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            consecutive_blank_rows = 0  # Reset consecutive_blank_rows counter
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table
    
    

def connectors(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = True

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = True
                    row_data.append("**")  # Start bold for the entire row
            else:
                value = str(cell.value)
                # Use regular expression to find and replace "mm2" with "mm<sup>2</sup>"
                value = re.sub(r'\bmm2\b', 'mm<sup>2</sup>', value)
                if make_bold:
                    row_data.append("**" + value + "**")
                else:
                    row_data.append(value)
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            if make_bold:
                make_bold = False  # Reset bold flag after first row
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table


  
def save_markdown_table(markdown_table, output_file):
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(markdown_table)

#####-------------------------------------------------------#####
# Custom automation script for generating all MDs automatically #
#####-------------------------------------------------------#####
def save_markdown_from_sheets():
    for filename in os.listdir("."):
        if filename.endswith(".xlsx"):
            wb = load_workbook(filename=filename, data_only=True)
            for sheet_name in wb.sheetnames:
                # Skip sheets named "rozcestnik" and "conns"
                if sheet_name in ["rozcestnik", "conns"]:
                    continue

                # Determine the function and output path based on the file pattern
                if filename.startswith("connectors"):
                    # For connectors files, use the connectors function
                    markdown_table = connectors(filename, sheet_name, "A1", "AA100")
                    # Check if the file is .en.xlsx and generate .en.md
                    if filename.endswith(".en.xlsx"):
                        output_file = f"../../source/md/{sheet_name}.en.md"
                    else:
                        output_file = f"../../source/md/{sheet_name}.md"
                elif sheet_name.startswith("commonHW"):
                    # For commonHW sheets, use nonBold function
                    markdown_table = nonBold(filename, sheet_name, "A1", "AA100")
                    output_file = f"../../source/md/{sheet_name}.md"
                else:
                    # For other files and sheets, use excel_to_markdown_table function
                    markdown_table = excel_to_markdown_table(filename, sheet_name, "A1", "AA100")

                    # Determine the base path based on sheet name prefixes
                    if sheet_name.startswith(("TGM", "TGZcontrolR", "TGZpMotion")):
                        base_path = "TGM"
                    elif sheet_name.startswith("TGS"):
                        base_path = "TGS"
                    else:
                        base_path = "TGZ"

                    # Determine if file is in English by checking for ".en.xlsx"
                    if filename.endswith(".en.xlsx"):
                        output_file = f"../../CZ/{base_path}/{sheet_name}/md/{filename.replace('.en.xlsx', '.en.md')}"
                    else:
                        output_file = f"../../CZ/{base_path}/{sheet_name}/md/{filename.replace('.xlsx', '.md')}"

                # Create necessary folders if they don't exist
                os.makedirs(os.path.dirname(output_file), exist_ok=True)

                # Save the markdown table
                save_markdown_table(markdown_table, output_file)

# Call the function
save_markdown_from_sheets()
