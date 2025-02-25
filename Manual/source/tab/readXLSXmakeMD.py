from openpyxl import load_workbook
from markdown import markdown
import re

def excel_to_markdown_table(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = True

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = True
            else:
                row_data.append(cell.value)
                if make_bold:
                    row_data[0] = "**" + row_data[0] + "**"
                    make_bold = False  # Reset bold flag after making first cell bold
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            consecutive_blank_rows = 0  # Reset consecutive_blank_rows counter
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table
    

def nonBold(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = False

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = False
            else:
                row_data.append(cell.value)
                if make_bold:
                    row_data[0] = "**" + row_data[0] + "**"
                    make_bold = False  # Reset bold flag after making first cell bold
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            consecutive_blank_rows = 0  # Reset consecutive_blank_rows counter
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table
    
    

def connectors(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = True

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = True
                    row_data.append("**")  # Start bold for the entire row
            else:
                value = str(cell.value)
                # Use regular expression to find and replace "mm2" with "mm<sup>2</sup>"
                value = re.sub(r'\bmm2\b', 'mm<sup>2</sup>', value)
                if make_bold:
                    row_data.append("**" + value + "**")
                else:
                    row_data.append(value)
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            if make_bold:
                make_bold = False  # Reset bold flag after first row
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table


  
def save_markdown_table(markdown_table, output_file):
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(markdown_table) 

# TGZ-D-48-13_26
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-48-13_26"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-48-13_26/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-50_100
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-50_100/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-50_100RI
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-50_100RI"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-50_100RI/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-48-50_100
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-48-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-48-50_100/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_250
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_250"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_250/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_250-O
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_250-O"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_250-O/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_250RI
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_250RI"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_250RI/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_300
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_300"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_300/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_300-O
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_300-O"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_300-O/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_300RI
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_300RI"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_300RI/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_425
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_425"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_425/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_425-O
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-100_425-O"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_425-O/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-230-5_15
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-230-5_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-230-5_15/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-320-5_10
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-320-5_10"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-320-5_10/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-320-5_15
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-320-5_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-320-5_15/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-3_9
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-400-3_9"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-3_9/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-7_15
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-400-7_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-7_15/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-10_20
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-400-10_20"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-10_20/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-14_30
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-400-14_30"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-14_30/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-3_9
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-560-3_9"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-3_9/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-7_15
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-560-7_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-7_15/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-10_20
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-560-10_20"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-10_20/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-30_50
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-560-30_50"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-30_50/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGS-320-10_15
file_path = "parameters.xlsx"
sheet_name = "TGS-320-10_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGS/TGS-320-10_15/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGS-560-25_50
file_path = "parameters.xlsx"
sheet_name = "TGS-560-25_50"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGS/TGS-560-25_50/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGS-560-50_100
file_path = "parameters.xlsx"
sheet_name = "TGS-560-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGS/TGS-560-50_100/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGMmini
file_path = "parameters.xlsx"
sheet_name = "TGMmini"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGMmini/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGMcontroller
file_path = "parameters.xlsx"
sheet_name = "TGMcontroller"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGMcontroller/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGMcontroller
file_path = "parameters.xlsx"
sheet_name = "TGZpMotion"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGZpMotion/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGMcontroller
file_path = "parameters.xlsx"
sheet_name = "TGZpMotion"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGZpMotion/md/parameters.md"
save_markdown_table(markdown_table, output_file)

##########################################################################################
# Other paramaters excel->md export
##########################################################################################

# Common DI params
file_path = "parameters.xlsx"
sheet_name = "commonHW_DI"
range_start = "A1"
range_end = "T100"
markdown_table = nonBold(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_commonHW_DI_tab.md"
save_markdown_table(markdown_table, output_file)

# Common DO params
file_path = "parameters.xlsx"
sheet_name = "commonHW_DO"
range_start = "A1"
range_end = "T100"
markdown_table = nonBold(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_commonHW_DO_tab.md"
save_markdown_table(markdown_table, output_file)

# Common AI params
file_path = "parameters.xlsx"
sheet_name = "commonHW_AI"
range_start = "A1"
range_end = "T100"
markdown_table = nonBold(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_commonHW_AI_tab.md"
save_markdown_table(markdown_table, output_file)

############################################ EN params ###################################

# TGZ-D-48-13_26
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-48-13_26"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-48-13_26/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-50_100
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-50_100/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-50_100RI
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-50_100RI"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-50_100RI/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-48-50_100
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-48-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-48-50_100/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_250
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_250"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_250/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_250-O
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_250-O"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_250-O/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_250RI
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_250RI"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_250RI/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_300
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_300"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_300/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_300-O
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_300-O"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_300-O/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_300RI
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_300RI"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_300RI/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_425
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_425"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_425/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-100_425-O
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-48-100_425-O"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-48-100_425-O/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-230-5_15
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-230-5_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-230-5_15/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-320-5_10
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-320-5_10"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-320-5_10/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-320-5_15
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-320-5_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-320-5_15/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-3_9
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-400-3_9"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-3_9/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-7_15
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-400-7_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-7_15/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-10_20
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-400-10_20"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-10_20/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-400-14_30
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-S-400-14_30"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-S-400-14_30/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-3_9
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-560-3_9"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-3_9/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-7_15
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-560-7_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-7_15/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-10_20
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-560-10_20"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-10_20/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-560-30_50
file_path = "parameters.en.xlsx"
sheet_name = "TGZ-D-560-30_50"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGZ/TGZ-D-560-30_50/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGS-320-10_15
file_path = "parameters.en.xlsx"
sheet_name = "TGS-320-10_15"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGS/TGS-320-10_15/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGS-560-25_50
file_path = "parameters.en.xlsx"
sheet_name = "TGS-560-25_50"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGS/TGS-560-25_50/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGS-560-50_100
file_path = "parameters.en.xlsx"
sheet_name = "TGS-560-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGS/TGS-560-50_100/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGMmini
file_path = "parameters.en.xlsx"
sheet_name = "TGMmini"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGMmini/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGMcontroller
file_path = "parameters.en.xlsx"
sheet_name = "TGMcontroller"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGMcontroller/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGMcontroller
file_path = "parameters.en.xlsx"
sheet_name = "TGZpMotion"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGZpMotion/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

# TGMcontroller
file_path = "parameters.en.xlsx"
sheet_name = "TGZpMotion"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../CZ/TGM/TGZpMotion/md/parameters.en.md"
save_markdown_table(markdown_table, output_file)

##########################################################################################
# Other paramaters excel->md export
##########################################################################################

# Common DI params
file_path = "parameters.en.xlsx"
sheet_name = "commonHW_DI"
range_start = "A1"
range_end = "T100"
markdown_table = nonBold(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_commonHW_DI_tab.en.md"
save_markdown_table(markdown_table, output_file)

# Common DO params
file_path = "parameters.en.xlsx"
sheet_name = "commonHW_DO"
range_start = "A1"
range_end = "T100"
markdown_table = nonBold(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_commonHW_DO_tab.en.md"
save_markdown_table(markdown_table, output_file)

# Common AI params
file_path = "parameters.en.xlsx"
sheet_name = "commonHW_AI"
range_start = "A1"
range_end = "T100"
markdown_table = nonBold(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_commonHW_AI_tab.en.md"
save_markdown_table(markdown_table, output_file)

##########################################################################################
# Connectors parameter MD generator
##########################################################################################

# X1 - 5pin weidmuller - BCZ 3.81/05/180 SN OR BX - +24V logic power
file_path = "connectors.xlsx"
sheet_name = "X1_24V_5pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# X2 - 2pin Phoenix - PC 5/ 2-STCL1-7,62 - DCbus +48V
file_path = "connectors.xlsx"
sheet_name = "X2_48_DC_1778065"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_48_DC_1778065.md"
save_markdown_table(markdown_table, output_file)

# X8 - 22pin Weidmuller - B2CF 3.50/22/180 SN OR BX - standard TGZ IO
file_path = "connectors.xlsx"
sheet_name = "X8_IO_22pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_IO_22pin_B2CF.md"
save_markdown_table(markdown_table, output_file)

# X10 - 4pin Weidmuller - B2CF 3.50/04/180 SN OR BX - standard TGZ CAN
file_path = "connectors.xlsx"
sheet_name = "X10_CAN_4pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X10_CAN_4pin_B2CF.md"
save_markdown_table(markdown_table, output_file)

# LED sig. - standard TGZ Green-Red Leds
file_path = "connectors.xlsx"
sheet_name = "LEDsigAx12"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/LEDsigAx12.md"
save_markdown_table(markdown_table, output_file)

# X5 - 12pin Weidmuller - B2CF 3.50/12/180 SN OR BX - standard TGZ FBE
file_path = "connectors.xlsx"
sheet_name = "X5_FBE_12pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_FBE_12pin_B2CF.md"
save_markdown_table(markdown_table, output_file)

# X6 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - standard TGZ FB1
file_path = "connectors.xlsx"
sheet_name = "X6_FB1_8pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X6_FB1_8pin_B2CF.md"
save_markdown_table(markdown_table, output_file)

# X7 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - standard TGZ FB2
file_path = "connectors.xlsx"
sheet_name = "X7_FB2_8pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X7_FB2_8pin_B2CF.md"
save_markdown_table(markdown_table, output_file)

# X3 - 6pin Weidmuller -  BLZP 5.08HC/06/180 SN OR BX - standard TGZ M1
file_path = "connectors.xlsx"
sheet_name = "X3_M1_6pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_6pin_BLZP.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller -  BLZP 5.08HC/06/180 SN OR BX - standard TGZ M2
file_path = "connectors.xlsx"
sheet_name = "X4_M2_6pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_6pin_BLZP.md"
save_markdown_table(markdown_table, output_file)

# X1 - 5pin Molex - 5055700501 - +24V logic power
file_path = "connectors.xlsx"
sheet_name = "X1_24V_5pin_Microlock"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_Microlock.md"
save_markdown_table(markdown_table, output_file)

# X3 - 2x Pressfit M5 - TGZ-S-48 DCbus
file_path = "connectors.xlsx"
sheet_name = "X3_DCbus_2pin_pressfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_DCbus_2pin_pressfit.md"
save_markdown_table(markdown_table, output_file)

# X3 - 2x Pressfit M5 - TGZ-S-48 M1
file_path = "connectors.xlsx"
sheet_name = "X3_M1_3pin_pressfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_3pin_pressfit.md"
save_markdown_table(markdown_table, output_file)

# X4 - 4pin Molex - 5055700401 - BR
file_path = "connectors.xlsx"
sheet_name = "X4_BR_4pin_Microlock"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_BR_4pin_Microlock.md"
save_markdown_table(markdown_table, output_file)

# X2 - 3pin Wago cage-clamp - 2636-1103 - DCbus
file_path = "connectors.xlsx"
sheet_name = "X2_DCbus_3pin_wago_2636"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_DCbus_3pin_wago_2636.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Wago cage-clamp - 2626-1104 - M1
file_path = "connectors.xlsx"
sheet_name = "X3_M1_4pin_wago_2626"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_4pin_wago_2626.md"
save_markdown_table(markdown_table, output_file)

# X4 - 4pin Wago cage-clamp - 2626-1104 - M2
file_path = "connectors.xlsx"
sheet_name = "X4_M2_4pin_wago_2626"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_4pin_wago_2626.md"
save_markdown_table(markdown_table, output_file)

# XBR - 6pin Weidmuller -  BLF 5.00HC/06/180F SN OR BX - Brake M1/M2
file_path = "connectors.xlsx"
sheet_name = "XBR_BR_6pin_BLF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/XBR_BR_6pin_BLF.md"
save_markdown_table(markdown_table, output_file)

# X1 - 5pin Molex Microfit 3.0 - 436450500 - +24V logic power
file_path = "connectors.xlsx"
sheet_name = "X1_24V_5pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_Microfit.md"
save_markdown_table(markdown_table, output_file)

# P7 - 4pin Molex Microfit 3.0 - 430250400 - Brake power + diag
file_path = "connectors.xlsx"
sheet_name = "P7_BR_4pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/P7_BR_4pin_Microfit.md"
save_markdown_table(markdown_table, output_file)

# P8 - 4pin Molex Microfit 3.0 - 430250400 - GND brake
file_path = "connectors.xlsx"
sheet_name = "P8_BR_4pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/P8_BR_4pin_Microfit.md"
save_markdown_table(markdown_table, output_file)

# P3 - 2pin Molex Microfit 3.0 - 436450200 - Ext. Therm.
file_path = "connectors.xlsx"
sheet_name = "P3_Term_2pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/P3_Term_2pin_Microfit.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller -  SLS 5.08/06/180FI SN OR BX - standard TGZ-S M
file_path = "connectors.xlsx"
sheet_name = "X4_M1_6pin_SLS"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M1_6pin_SLS.md"
save_markdown_table(markdown_table, output_file)

# X2 - 10pin Weidmuller -  BLZP 5.08HC/10/180 SN OR BX - standard TGZ-S PWR
file_path = "connectors.xlsx"
sheet_name = "X2_PWR_10pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_PWR_10pin_BLZP.md"
save_markdown_table(markdown_table, output_file)

# X2 - 3pin Phoenix - PC 5/ 3-STCL1-7,62 - DCbus 320V
file_path = "connectors.xlsx"
sheet_name = "X2_320_DC_1778078"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_320_DC_1778078.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller -  BLF 7.62HP/06/180F - standard TGZ-S-400 M1
file_path = "connectors.xlsx"
sheet_name = "X4_M1_6pin_BLF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M1_6pin_BLF.md"
save_markdown_table(markdown_table, output_file)

# X2 - 12pin Weidmuller -  BLZ 7.62HP/12/180F - standard TGZ-S-400 PWR
file_path = "connectors.xlsx"
sheet_name = "X2_PWR_12pin_BLZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_PWR_12pin_BLZ.md"
save_markdown_table(markdown_table, output_file)

# X2 - 2x M8 - TGZ-D-560-30_50 DCbus
file_path = "connectors.xlsx"
sheet_name = "X2_D560DCbus"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_D560DCbus.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Wago 2636 - TGZ-D-560-30_50 M1
file_path = "connectors.xlsx"
sheet_name = "X3_M1_4pin_wago_2636"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_4pin_wago_2636.md"
save_markdown_table(markdown_table, output_file)

# X4 - 4pin Wago 2636 - TGZ-D-560-30_50 M2
file_path = "connectors.xlsx"
sheet_name = "X4_M2_4pin_wago_2636"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_4pin_wago_2636.md"
save_markdown_table(markdown_table, output_file)

# X14 - 4pin Weidmuller - LSF-SMT 5.00/04/90 3.5SN BK RL - TGZ-D-560-30_50 BR1
file_path = "connectors.xlsx"
sheet_name = "X14_BR1_4pin_LSF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X14_BR1_4pin_LSF.md"
save_markdown_table(markdown_table, output_file)

# X15 - 4pin Weidmuller - LSF-SMT 5.00/04/90 3.5SN BK RL - TGZ-D-560-30_50 BR2
file_path = "connectors.xlsx"
sheet_name = "X15_BR2_4pin_LSF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X15_BR2_4pin_LSF.md"
save_markdown_table(markdown_table, output_file)

# X1 - 3pin Phoenix PC 5/ 3-STCL1-7,62 - ACIN
file_path = "connectors.xlsx"
sheet_name = "X1_ACIN_PC5"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_ACIN_PC5.md"
save_markdown_table(markdown_table, output_file)

# X2 - 8pin Phoenix PC 5/ 8-STCL1-7,62 - DC bus
file_path = "connectors.xlsx"
sheet_name = "X2_DC_8pin_PC5"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_DC_8pin_PC5.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Weidmuller BLF 2.50/04/180 SN BK BX
file_path = "connectors.xlsx"
sheet_name = "X3_24V_BLF_2_5"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_24V_BLF_2_5.md"
save_markdown_table(markdown_table, output_file)

# X5 - 10pin Weidmuller B2CF 3.50/10/180 SN OR BX - Mini DO
file_path = "connectors.xlsx"
sheet_name = "X5_DI_10pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_DI_10pin_B2CF.md"
save_markdown_table(markdown_table, output_file)

# X10 - 10pin Weidmuller B2CF 3.50/10/180 SN OR BX - Mini DI
file_path = "connectors.xlsx"
sheet_name = "X10_DO_10pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X10_DO_10pin_B2CF.md"
save_markdown_table(markdown_table, output_file)

# S1 - Dip switch Mini CAN term
file_path = "connectors.xlsx"
sheet_name = "S1_SWITCH_CAN"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/S1_SWITCH_CAN.md"
save_markdown_table(markdown_table, output_file)

# X1 - 5pin weidmuller - BCZ 3.81/05/180 SN OR BX - +24V logic power TGMcontroller
file_path = "connectors.xlsx"
sheet_name = "X1_24V_5pin_BCZ_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_BCZ_TGM.md"
save_markdown_table(markdown_table, output_file)

# X5 - 12pin Weidmuller - B2CF 3.50/12/180 SN OR BX - TGM FBE
file_path = "connectors.xlsx"
sheet_name = "X5_FBE_12pin_B2CF_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_FBE_12pin_B2CF_TGM.md"
save_markdown_table(markdown_table, output_file)

# X6 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - TGM FB1
file_path = "connectors.xlsx"
sheet_name = "X6_FB1_8pin_B2CF_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X6_FB1_8pin_B2CF_TGM.md"
save_markdown_table(markdown_table, output_file)

# X7 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - TGM FB2
file_path = "connectors.xlsx"
sheet_name = "X7_FB2_8pin_B2CF_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X7_FB2_8pin_B2CF_TGM.md"
save_markdown_table(markdown_table, output_file)

# X4 - 12pin Molex ClikMate - 5031491200 - TGZ RI FBE
file_path = "connectors.xlsx"
sheet_name = "X4_FBE_12pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_FBE_12pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X5 - 10pin Molex ClikMate - 5031491000 - TGZ RI FB1
file_path = "connectors.xlsx"
sheet_name = "X5_FB1_10pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_FB1_10pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X6 - 10pin Molex ClikMate - 5031491000 - TGZ RI FB2
file_path = "connectors.xlsx"
sheet_name = "X6_FB2_10pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X6_FB2_10pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X7 - 12pin Molex ClikMate - 5031491200 - TGZ RI AIN, DITTL
file_path = "connectors.xlsx"
sheet_name = "X7_AIN_12pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X7_AIN_12pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X8 - 18pin Molex ClikMate - 5031491800 - TGZ RI DI
file_path = "connectors.xlsx"
sheet_name = "X8_DIO_18pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_DIO_18pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X10 - 8pin Molex ClikMate - 5031490800 - TGZ RI CAN
file_path = "connectors.xlsx"
sheet_name = "X10_CAN_8pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X10_CAN_8pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X11 - 10pin Molex ClikMate - 5031491000 - TGZ RI FB3
file_path = "connectors.xlsx"
sheet_name = "X11_FB3_10pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X11_FB3_10pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X12 - 8pin Molex ClikMate - 5031490800 - TGZ RI ENET, ECAT
file_path = "connectors.xlsx"
sheet_name = "X12_UDP_8pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X12_UDP_8pin_ClikMate.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Weidmuller BCZ - TGS320 DO
file_path = "connectors.xlsx"
sheet_name = "X3_DO_4pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_DO_4pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# X2 - 3pin Weidmuller BLZ 7.62 - TGZ560-3/9 DCBUS
file_path = "connectors.xlsx"
sheet_name = "X2_560_DC_3pin_BLZ__7_62"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_560_DC_3pin_BLZ__7_62.md"
save_markdown_table(markdown_table, output_file)

# X3 - 6pin Weidmuller BLZ 7.62 - TGZ560-3/9 M1
file_path = "connectors.xlsx"
sheet_name = "X3_M1_6pin_BLZ__7_62"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_6pin_BLZ__7_62.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller BLZ 7.62 - TGZ560-3/9 M2
file_path = "connectors.xlsx"
sheet_name = "X4_M2_6pin_BLZ__7_62"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_6pin_BLZ__7_62.md"
save_markdown_table(markdown_table, output_file)

# X4 - 5pin Weidmuller - BCZ 3.81/05/180 SN OR BX - +24V logic power TGS560
file_path = "connectors.xlsx"
sheet_name = "X4_TGS560_24V_5pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_TGS560_24V_5pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# S1 - DIP switch 4pos. TGS560
file_path = "connectors.xlsx"
sheet_name = "S1_TGS560_DIP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/S1_TGS560_DIP.md"
save_markdown_table(markdown_table, output_file)

# S1 - DIP switch 4pos. TGS560
file_path = "connectors.xlsx"
sheet_name = "LED_TGS560"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/LED_TGS560.md"
save_markdown_table(markdown_table, output_file)

# X4 - 4pin Weidmuller - BVF 7.62HP/04/180F - ACIN TGS560-25
file_path = "connectors.xlsx"
sheet_name = "X4_ACIN_4pin_TGS560_25"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_ACIN_4pin_TGS560_25.md"
save_markdown_table(markdown_table, output_file)

# X1 - 6pin Weidmuller - BVF 7.62HP/06/180F - DCout TGS560-25
file_path = "connectors.xlsx"
sheet_name = "X1_DC_6pin_TGS560_25"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_DC_6pin_TGS560_25.md"
save_markdown_table(markdown_table, output_file)

# X5 - 3pin Weidmuller - BLZ 7.62HP/03/180F - BRres TGS560-25
file_path = "connectors.xlsx"
sheet_name = "X5_RBR_3pin_TGS560_25"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_RBR_3pin_TGS560_25.md"
save_markdown_table(markdown_table, output_file)


##########################################################################################
# Connectors EN parameter MD generator
##########################################################################################

# X1 - 5pin weidmuller - BCZ 3.81/05/180 SN OR BX - +24V logic power
file_path = "connectors.en.xlsx"
sheet_name = "X1_24V_5pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_BCZ.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 2pin Phoenix - PC 5/ 2-STCL1-7,62 - DCbus +48V
file_path = "connectors.en.xlsx"
sheet_name = "X2_48_DC_1778065"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_48_DC_1778065.en.md"
save_markdown_table(markdown_table, output_file)

# X8 - 22pin Weidmuller - B2CF 3.50/22/180 SN OR BX - standard TGZ IO
file_path = "connectors.en.xlsx"
sheet_name = "X8_IO_22pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_IO_22pin_B2CF.en.md"
save_markdown_table(markdown_table, output_file)

# X10 - 4pin Weidmuller - B2CF 3.50/04/180 SN OR BX - standard TGZ CAN
file_path = "connectors.en.xlsx"
sheet_name = "X10_CAN_4pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X10_CAN_4pin_B2CF.en.md"
save_markdown_table(markdown_table, output_file)

# LED sig. - standard TGZ Green-Red Leds
file_path = "connectors.en.xlsx"
sheet_name = "LEDsigAx12"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/LEDsigAx12.en.md"
save_markdown_table(markdown_table, output_file)

# X5 - 12pin Weidmuller - B2CF 3.50/12/180 SN OR BX - standard TGZ FBE
file_path = "connectors.en.xlsx"
sheet_name = "X5_FBE_12pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_FBE_12pin_B2CF.en.md"
save_markdown_table(markdown_table, output_file)

# X6 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - standard TGZ FB1
file_path = "connectors.en.xlsx"
sheet_name = "X6_FB1_8pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X6_FB1_8pin_B2CF.en.md"
save_markdown_table(markdown_table, output_file)

# X7 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - standard TGZ FB2
file_path = "connectors.en.xlsx"
sheet_name = "X7_FB2_8pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X7_FB2_8pin_B2CF.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 6pin Weidmuller -  BLZP 5.08HC/06/180 SN OR BX - standard TGZ M1
file_path = "connectors.en.xlsx"
sheet_name = "X3_M1_6pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_6pin_BLZP.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller -  BLZP 5.08HC/06/180 SN OR BX - standard TGZ M2
file_path = "connectors.en.xlsx"
sheet_name = "X4_M2_6pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_6pin_BLZP.en.md"
save_markdown_table(markdown_table, output_file)

# X1 - 5pin Molex - 5055700501 - +24V logic power
file_path = "connectors.en.xlsx"
sheet_name = "X1_24V_5pin_Microlock"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_Microlock.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 2x Pressfit M5 - TGZ-S-48 DCbus
file_path = "connectors.en.xlsx"
sheet_name = "X3_DCbus_2pin_pressfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_DCbus_2pin_pressfit.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 2x Pressfit M5 - TGZ-S-48 M1
file_path = "connectors.en.xlsx"
sheet_name = "X3_M1_3pin_pressfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_3pin_pressfit.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 4pin Molex - 5055700401 - BR
file_path = "connectors.en.xlsx"
sheet_name = "X4_BR_4pin_Microlock"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_BR_4pin_Microlock.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 3pin Wago cage-clamp - 2636-1103 - DCbus
file_path = "connectors.en.xlsx"
sheet_name = "X2_DCbus_3pin_wago_2636"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_DCbus_3pin_wago_2636.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Wago cage-clamp - 2626-1104 - M1
file_path = "connectors.en.xlsx"
sheet_name = "X3_M1_4pin_wago_2626"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_4pin_wago_2626.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 4pin Wago cage-clamp - 2626-1104 - M2
file_path = "connectors.en.xlsx"
sheet_name = "X4_M2_4pin_wago_2626"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_4pin_wago_2626.en.md"
save_markdown_table(markdown_table, output_file)

# XBR - 6pin Weidmuller -  BLF 5.00HC/06/180F SN OR BX - Brake M1/M2
file_path = "connectors.en.xlsx"
sheet_name = "XBR_BR_6pin_BLF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/XBR_BR_6pin_BLF.en.md"
save_markdown_table(markdown_table, output_file)

# X1 - 5pin Molex Microfit 3.0 - 436450500 - +24V logic power
file_path = "connectors.en.xlsx"
sheet_name = "X1_24V_5pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_Microfit.en.md"
save_markdown_table(markdown_table, output_file)

# P7 - 4pin Molex Microfit 3.0 - 430250400 - Brake power + diag
file_path = "connectors.en.xlsx"
sheet_name = "P7_BR_4pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/P7_BR_4pin_Microfit.en.md"
save_markdown_table(markdown_table, output_file)

# P8 - 4pin Molex Microfit 3.0 - 430250400 - GND brake
file_path = "connectors.en.xlsx"
sheet_name = "P8_BR_4pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/P8_BR_4pin_Microfit.en.md"
save_markdown_table(markdown_table, output_file)

# P3 - 2pin Molex Microfit 3.0 - 436450200 - Ext. Therm.
file_path = "connectors.en.xlsx"
sheet_name = "P3_Term_2pin_Microfit"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/P3_Term_2pin_Microfit.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller -  SLS 5.08/06/180FI SN OR BX - standard TGZ-S M
file_path = "connectors.en.xlsx"
sheet_name = "X4_M1_6pin_SLS"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M1_6pin_SLS.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 10pin Weidmuller -  BLZP 5.08HC/10/180 SN OR BX - standard TGZ-S PWR
file_path = "connectors.en.xlsx"
sheet_name = "X2_PWR_10pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_PWR_10pin_BLZP.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 3pin Phoenix - PC 5/ 3-STCL1-7,62 - DCbus 320V
file_path = "connectors.en.xlsx"
sheet_name = "X2_320_DC_1778078"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_320_DC_1778078.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller -  BLF 7.62HP/06/180F - standard TGZ-S-400 M1
file_path = "connectors.en.xlsx"
sheet_name = "X4_M1_6pin_BLF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M1_6pin_BLF.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 12pin Weidmuller -  BLZ 7.62HP/12/180F - standard TGZ-S-400 PWR
file_path = "connectors.en.xlsx"
sheet_name = "X2_PWR_12pin_BLZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_PWR_12pin_BLZ.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 2x M8 - TGZ-D-560-30_50 DCbus
file_path = "connectors.en.xlsx"
sheet_name = "X2_D560DCbus"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_D560DCbus.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Wago 2636 - TGZ-D-560-30_50 M1
file_path = "connectors.en.xlsx"
sheet_name = "X3_M1_4pin_wago_2636"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_4pin_wago_2636.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 4pin Wago 2636 - TGZ-D-560-30_50 M2
file_path = "connectors.en.xlsx"
sheet_name = "X4_M2_4pin_wago_2636"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_4pin_wago_2636.en.md"
save_markdown_table(markdown_table, output_file)

# X14 - 4pin Weidmuller - LSF-SMT 5.00/04/90 3.5SN BK RL - TGZ-D-560-30_50 BR1
file_path = "connectors.en.xlsx"
sheet_name = "X14_BR1_4pin_LSF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X14_BR1_4pin_LSF.en.md"
save_markdown_table(markdown_table, output_file)

# X15 - 4pin Weidmuller - LSF-SMT 5.00/04/90 3.5SN BK RL - TGZ-D-560-30_50 BR2
file_path = "connectors.en.xlsx"
sheet_name = "X15_BR2_4pin_LSF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X15_BR2_4pin_LSF.en.md"
save_markdown_table(markdown_table, output_file)

# X1 - 3pin Phoenix PC 5/ 3-STCL1-7,62 - ACIN
file_path = "connectors.en.xlsx"
sheet_name = "X1_ACIN_PC5"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_ACIN_PC5.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 8pin Phoenix PC 5/ 8-STCL1-7,62 - DC bus
file_path = "connectors.en.xlsx"
sheet_name = "X2_DC_8pin_PC5"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_DC_8pin_PC5.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Weidmuller BLF 2.50/04/180 SN BK BX
file_path = "connectors.en.xlsx"
sheet_name = "X3_24V_BLF_2_5"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_24V_BLF_2_5.en.md"
save_markdown_table(markdown_table, output_file)

# X5 - 10pin Weidmuller B2CF 3.50/10/180 SN OR BX - Mini DO
file_path = "connectors.en.xlsx"
sheet_name = "X5_DI_10pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_DI_10pin_B2CF.en.md"
save_markdown_table(markdown_table, output_file)

# X10 - 10pin Weidmuller B2CF 3.50/10/180 SN OR BX - Mini DI
file_path = "connectors.en.xlsx"
sheet_name = "X10_DO_10pin_B2CF"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X10_DO_10pin_B2CF.en.md"
save_markdown_table(markdown_table, output_file)

# S1 - Dip switch Mini CAN term
file_path = "connectors.en.xlsx"
sheet_name = "S1_SWITCH_CAN"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/S1_SWITCH_CAN.en.md"
save_markdown_table(markdown_table, output_file)

# X1 - 5pin weidmuller - BCZ 3.81/05/180 SN OR BX - +24V logic power TGMcontroller
file_path = "connectors.en.xlsx"
sheet_name = "X1_24V_5pin_BCZ_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X1_24V_5pin_BCZ_TGM.en.md"
save_markdown_table(markdown_table, output_file)

# X5 - 12pin Weidmuller - B2CF 3.50/12/180 SN OR BX - TGM FBE
file_path = "connectors.en.xlsx"
sheet_name = "X5_FBE_12pin_B2CF_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_FBE_12pin_B2CF_TGM.en.md"
save_markdown_table(markdown_table, output_file)

# X6 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - TGM FB1
file_path = "connectors.en.xlsx"
sheet_name = "X6_FB1_8pin_B2CF_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X6_FB1_8pin_B2CF_TGM.en.md"
save_markdown_table(markdown_table, output_file)

# X7 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - TGM FB2
file_path = "connectors.en.xlsx"
sheet_name = "X7_FB2_8pin_B2CF_TGM"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X7_FB2_8pin_B2CF_TGM.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 12pin Molex ClikMate - 5031491200 - TGZ RI FBE
file_path = "connectors.en.xlsx"
sheet_name = "X4_FBE_12pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_FBE_12pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X5 - 10pin Molex ClikMate - 5031491000 - TGZ RI FB1
file_path = "connectors.en.xlsx"
sheet_name = "X5_FB1_10pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X5_FB1_10pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X6 - 10pin Molex ClikMate - 5031491000 - TGZ RI FB2
file_path = "connectors.en.xlsx"
sheet_name = "X6_FB2_10pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X6_FB2_10pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X7 - 12pin Molex ClikMate - 5031491200 - TGZ RI AIN, DITTL
file_path = "connectors.en.xlsx"
sheet_name = "X7_AIN_12pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X7_AIN_12pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X8 - 18pin Molex ClikMate - 5031491800 - TGZ RI DI
file_path = "connectors.en.xlsx"
sheet_name = "X8_DIO_18pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X8_DIO_18pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X10 - 8pin Molex ClikMate - 5031490800 - TGZ RI CAN
file_path = "connectors.en.xlsx"
sheet_name = "X10_CAN_8pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X10_CAN_8pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X11 - 10pin Molex ClikMate - 5031491000 - TGZ RI FB3
file_path = "connectors.en.xlsx"
sheet_name = "X11_FB3_10pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X11_FB3_10pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X12 - 8pin Molex ClikMate - 5031490800 - TGZ RI ENET, ECAT
file_path = "connectors.en.xlsx"
sheet_name = "X12_UDP_8pin_ClikMate"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X12_UDP_8pin_ClikMate.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 4pin Weidmuller BCZ - TGS320 DO
file_path = "connectors.en.xlsx"
sheet_name = "X3_DO_4pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_DO_4pin_BCZ.en.md"
save_markdown_table(markdown_table, output_file)

# X2 - 3pin Weidmuller BLZ 7.62 - TGZ560-3/9 DCBUS
file_path = "connectors.en.xlsx"
sheet_name = "X2_560_DC_3pin_BLZ__7_62"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X2_560_DC_3pin_BLZ__7_62.en.md"
save_markdown_table(markdown_table, output_file)

# X3 - 6pin Weidmuller BLZ 7.62 - TGZ560-3/9 M1
file_path = "connectors.en.xlsx"
sheet_name = "X3_M1_6pin_BLZ__7_62"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X3_M1_6pin_BLZ__7_62.en.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller BLZ 7.62 - TGZ560-3/9 M2
file_path = "connectors.en.xlsx"
sheet_name = "X4_M2_6pin_BLZ__7_62"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../source/md/X4_M2_6pin_BLZ__7_62.en.md"
save_markdown_table(markdown_table, output_file)