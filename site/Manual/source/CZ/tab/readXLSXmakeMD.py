from openpyxl import load_workbook
from markdown import markdown
import re

def excel_to_markdown_table(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = True

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = True
            else:
                row_data.append(cell.value)
                if make_bold:
                    row_data[0] = "**" + row_data[0] + "**"
                    make_bold = False  # Reset bold flag after making first cell bold
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            consecutive_blank_rows = 0  # Reset consecutive_blank_rows counter
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table

def connectors(file_path, sheet_name, range_start, range_end):
    # Load the Excel workbook
    wb = load_workbook(filename=file_path, data_only=True)

    # Select the specified sheet
    sheet = wb[sheet_name]

    # Initialize list to store table rows
    rows = []

    # Flag to track if bold formatting is needed
    make_bold = True

    # Iterate over the specified range of cells and extract data
    consecutive_blank_rows = 0  # Counter for consecutive blank rows
    for row_idx, row in enumerate(sheet[range_start:range_end], start=1):
        row_data = []
        row_is_blank = True  # Flag to track if current row is blank
        for cell_idx, cell in enumerate(row, start=1):
            # Skip empty cells
            if cell.value in (None, "None"):
                if cell_idx == 1:
                    make_bold = True
                    row_data.append("**")  # Start bold for the entire row
            else:
                value = str(cell.value)
                # Use regular expression to find and replace "mm2" with "mm<sup>2</sup>"
                value = re.sub(r'\bmm2\b', 'mm<sup>2</sup>', value)
                if make_bold:
                    row_data.append("**" + value + "**")
                else:
                    row_data.append(value)
                row_is_blank = False  # Set row_is_blank to False if non-empty cell found
        if not row_is_blank:  # Append non-empty rows
            rows.append(row_data)
            if make_bold:
                make_bold = False  # Reset bold flag after first row
        else:
            consecutive_blank_rows += 1  # Increment consecutive_blank_rows counter
            if consecutive_blank_rows >= 2:  # Skip printing rows with two or more consecutive blank rows
                break

    # Convert data to Markdown table format
    markdown_table = ""
    for idx, row_data in enumerate(rows):
        if idx == 0:
            markdown_table += "| " + " | ".join(cell for cell in row_data) + " |\n"
            markdown_table += "| " + " | ".join(":---:" for _ in row_data) + " |\n"
        else:
            markdown_table += "| " + " | ".join(str(cell) for cell in row_data) + " |\n"

    return markdown_table


  
def save_markdown_table(markdown_table, output_file):
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(markdown_table)

# TGZ-D-48-13_26
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-48-13_26"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../../CZ/TGZ/TGZ-D-48-13/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-S-48-50_100
file_path = "parameters.xlsx"
sheet_name = "TGZ-S-48-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../../CZ/TGZ/TGZ-S-48-50_100/md/parameters.md"
save_markdown_table(markdown_table, output_file)

# TGZ-D-48-50_100
file_path = "parameters.xlsx"
sheet_name = "TGZ-D-48-50_100"
range_start = "A1"
range_end = "B100"
markdown_table = excel_to_markdown_table(file_path, sheet_name, range_start, range_end)
output_file = "../../../CZ/TGZ/TGZ-D-48-50_100/md/parameters.md"
save_markdown_table(markdown_table, output_file)

##########################################################################################
# Connectors parameter MD generator
##########################################################################################

# X1 - 5pin weidmuller - BCZ 3.81/05/180 SN OR BX - +24V logic power
file_path = "connectors.xlsx"
sheet_name = "X1_24V_5pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X1_24V_5pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# X2 - 2pin Phoenix - PC 5/ 2-STCL1-7,62 - DCbus +48V
file_path = "connectors.xlsx"
sheet_name = "X2_48_DC_1778065"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X2_48_DC_1778065.md"
save_markdown_table(markdown_table, output_file)

# X8 - 22pin Weidmuller - B2CF 3.50/22/180 SN OR BX - standard TGZ IO
file_path = "connectors.xlsx"
sheet_name = "X8_IO_22pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X8_IO_22pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# X10 - 4pin Weidmuller - B2CF 3.50/04/180 SN OR BX - standard TGZ CAN
file_path = "connectors.xlsx"
sheet_name = "X10_CAN_4pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X10_CAN_4pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# LED sig. - standard TGZ Green-Red Leds
file_path = "connectors.xlsx"
sheet_name = "LEDsigAx12"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/LEDsigAx12.md"
save_markdown_table(markdown_table, output_file)

# X5 - 12pin Weidmuller - B2CF 3.50/12/180 SN OR BX - standard TGZ FBE
file_path = "connectors.xlsx"
sheet_name = "X5_FBE_12pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X5_FBE_12pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# X6 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - standard TGZ FB1
file_path = "connectors.xlsx"
sheet_name = "X6_FB1_8pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X6_FB1_8pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# X7 - 8pin Weidmuller - B2CF 3.50/08/180 SN OR BX - standard TGZ FB2
file_path = "connectors.xlsx"
sheet_name = "X7_FB2_8pin_BCZ"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X7_FB2_8pin_BCZ.md"
save_markdown_table(markdown_table, output_file)

# X3 - 6pin Weidmuller -  BLZP 5.08HC/06/180 SN OR BX - standard TGZ M1
file_path = "connectors.xlsx"
sheet_name = "X3_M1_6pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X3_M1_6pin_BLZP.md"
save_markdown_table(markdown_table, output_file)

# X4 - 6pin Weidmuller -  BLZP 5.08HC/06/180 SN OR BX - standard TGZ M2
file_path = "connectors.xlsx"
sheet_name = "X4_M2_6pin_BLZP"
range_start = "A1"
range_end = "D100"
markdown_table = connectors(file_path, sheet_name, range_start, range_end)
output_file = "../../../source/CZ/md/X4_M2_6pin_BLZP.md"
save_markdown_table(markdown_table, output_file)